import streamlit as st
import pandas as pd
from io import BytesIO
import os
import re
from typing import List, Tuple, Dict

# -----------------------------
# Page configuration
# -----------------------------
st.set_page_config(
    layout="wide",
    page_title="Muuto Lookup Tool",
    page_icon="favicon.png",  # keep same asset name
)

# -----------------------------
# Styling reused from previous app
# -----------------------------
st.markdown(
    """
<style>
    .stApp, body { background-color: #EFEEEB !important; }
    .main .block-container { background-color: #EFEEEB !important; padding-top: 2rem; }
    h1, h2, h3 { text-transform: none !important; }
    h1 { color: #333; }
    h2 { color: #1E40AF; padding-bottom: 5px; margin-top: 30px; margin-bottom: 15px; }
    h3 { color: #1E40AF; font-size: 1.25em; padding-bottom: 3px; margin-top: 20px; margin-bottom: 10px; }
    h4 { color: #102A63; font-size: 1.1em; margin-top: 15px; margin-bottom: 5px; }

    div[data-testid="stAlert"] { background-color: #f0f2f6 !important; border: 1px solid #D1D5DB !important; border-radius: 0.25rem !important; }
    div[data-testid="stAlert"] > div:first-child { background-color: transparent !important; }
    div[data-testid="stAlert"] div[data-testid="stMarkdownContainer"], div[data-testid="stAlert"] div[data-testid="stMarkdownContainer"] p { color: #31333F !important; }
    div[data-testid="stAlert"] svg { fill: #4B5563 !important; }

    /* Inputs */
    div[data-testid="stTextArea"] textarea,
    div[data-testid="stTextInput"] input,
    div[data-testid="stSelectbox"] div[data-baseweb="select"] > div:first-child,
    div[data-testid="stMultiSelect"] div[data-baseweb="input"],
    div[data-testid="stMultiSelect"] > div > div[data-baseweb="select"] > div:first-child {
        background-color: #FFFFFF !important; color: #000000 !important; border: 1px solid #CCCCCC !important;
    }
    div[data-testid="stTextArea"] textarea:focus,
    div[data-testid="stTextInput"] input:focus,
    div[data-testid="stSelectbox"] div[data-baseweb="select"][aria-expanded="true"] > div:first-child,
    div[data-testid="stMultiSelect"] div[data-baseweb="input"]:focus-within,
    div[data-testid="stMultiSelect"] div[aria-expanded="true"] {
        border-color: #5B4A14 !important; box-shadow: 0 0 0 1px #5B4A14 !important;
    }

    /* Buttons */
    div[data-testid="stDownloadButton"] button[data-testid^="stBaseButton"],
    div[data-testid="stButton"] button[data-testid^="stBaseButton"] {
        border: 1px solid #5B4A14 !important; background-color: #FFFFFF !important; color: #5B4A14 !important;
        padding: 0.375rem 0.75rem !important; font-size: 1rem !important; line-height: 1.5 !important; border-radius: 0.25rem !important;
        transition: color 0.15s ease-in-out, background-color 0.15s ease-in-out, border-color 0.15s ease-in-out, box-shadow 0.15s ease-in-out !important; font-weight: 500 !important;
        text-transform: none !important;
    }
    div[data-testid="stDownloadButton"] button[data-testid^="stBaseButton"]:hover,
    div[data-testid="stButton"] button[data-testid^="stBaseButton"]:hover { background-color: #5B4A14 !important; color: #FFFFFF !important; border-color: #5B4A14 !important; }
    div[data-testid="stDownloadButton"] button[data-testid^="stBaseButton"]:active,
    div[data-testid="stDownloadButton"] button[data-testid^="stBaseButton"]:focus,
    div[data-testid="stButton"] button[data-testid^="stBaseButton"]:active,
    div[data-testid="stButton"] button[data-testid^="stBaseButton"]:focus { background-color: #4A3D10 !important; color: #FFFFFF !important; border-color: #4A3D10 !important; box-shadow: 0 0 0 0.2rem rgba(91, 74, 20, 0.4) !important; outline: none !important; }
</style>
""",
    unsafe_allow_html=True,
)

# -----------------------------
# Constants
# -----------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_MAPPING_PATH = os.path.join(BASE_DIR, "mapping.xlsx")
LOGO_PATH = os.path.join(BASE_DIR, "muuto_logo.png")

# Expected output order and header names
OUTPUT_HEADERS = [
    "New Item No.",
    "OLD Item-variant",
    "Ean no.",  # use exact header spelling from mapping file
    "Description",
    "Family",
    "Category",
]

# -----------------------------
# Helpers
# -----------------------------

def parse_pasted_ids(raw: str) -> List[str]:
    """Split on whitespace, commas, semicolons, and tabs. Preserve hyphens and leading zeros."""
    if not raw:
        return []
    tokens = re.split(r"[\s,;]+", raw.strip())
    # filter empties and normalize by stripping outer quotes
    cleaned = [t.strip().strip('"').strip("'") for t in tokens if t.strip()]
    # keep distinct while preserving order
    seen = set()
    result = []
    for t in cleaned:
        if t not in seen:
            seen.add(t)
            result.append(t)
    return result


def _read_mapping_with_values(xlsx_bytes: bytes) -> pd.DataFrame:
    """Read Excel with formula *values* using openpyxl data_only=True, then build DataFrame.
    Ensures concatenation results in column 'OLD Item-variant' are resolved.
    """
    from openpyxl import load_workbook
    wb = load_workbook(filename=BytesIO(xlsx_bytes), data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = rows[1:]
    df = pd.DataFrame(data, columns=headers)
    return df


def load_mapping(file_uploader) -> Tuple[pd.DataFrame, str]:
    if file_uploader is not None:
        xlsx_bytes = file_uploader.getvalue()
        try:
            df = _read_mapping_with_values(xlsx_bytes)
            source = "uploaded"
            return df, source
        except Exception as e:
            st.error(f"Kunne ikke læse den uploadede mapping-fil: {e}")
            return pd.DataFrame(), "error"
    # fallback to bundled file path
    if os.path.exists(DEFAULT_MAPPING_PATH):
        try:
            with open(DEFAULT_MAPPING_PATH, "rb") as f:
                df = _read_mapping_with_values(f.read())
            source = "bundled"
            return df, source
        except Exception as e:
            st.error(f"Kunne ikke læse mapping.xlsx i repoet: {e}")
            return pd.DataFrame(), "error"
    st.error("Ingen mapping-fil fundet. Upload en mapping.xlsx.")
    return pd.DataFrame(), "missing"


def normalize_columns_case_insensitive(df: pd.DataFrame, required: List[str]) -> Dict[str, str]:
    """Return a dict mapping canonical required names to actual df columns matched case-insensitively."""
    lower_map = {c.lower(): c for c in df.columns}
    out = {}
    for name in required:
        key = name.lower()
        if key in lower_map:
            out[name] = lower_map[key]
        else:
            out[name] = None
    return out


def select_and_order(df: pd.DataFrame, colmap: Dict[str, str]) -> pd.DataFrame:
    cols = []
    for cname in OUTPUT_HEADERS:
        actual = colmap.get(cname)
        if actual is None:
            df[cname] = None
            cols.append(cname)
        else:
            cols.append(actual)
    out = df[cols].copy()
    # Rename to canonical output headers if the actual names differ
    rename_map = {colmap[h]: h for h in OUTPUT_HEADERS if colmap.get(h) and colmap[h] != h}
    if rename_map:
        out = out.rename(columns=rename_map)
    return out


def to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Lookup Output") -> bytes:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    return buf.getvalue()

# -----------------------------
# Header with logo
# -----------------------------
left, right = st.columns([6, 1])
with left:
    st.title("Muuto Mapping Lookup")
    st.caption("Indsæt Muuto vare-variant numre eller EAN numre. Vi matcher mod mapping.xlsx og returnerer et Excel-udtræk.")
with right:
    if os.path.exists(LOGO_PATH):
        st.image(LOGO_PATH, width=120)

st.markdown("---")

# -----------------------------
# Controls
# -----------------------------
col_a, col_b = st.columns([2, 1])
with col_a:
    uploaded = st.file_uploader("Upload mapping.xlsx (valgfrit)", type=["xlsx"], help="Hvis du ikke uploader, bruger appen mapping.xlsx fra repoet.")
with col_b:
    st.write(" ")
    st.write(" ")
    sample_btn = st.button("Indsæt eksempelværdier")

if sample_btn:
    st.session_state["paste_box"] = """5710562801234\nMTO-CHAIR-001-01\n5710562805678\nMTO-SOFA-CHAIS-LEFT-22"""

raw_input = st.text_area(
    "Indsæt liste her",
    key="paste_box",
    height=200,
    placeholder="Indsæt numre adskilt af linjeskift, komma eller semikolon",
)

ids = parse_pasted_ids(raw_input)

mapping_df, source = load_mapping(uploaded)

required_headers = OUTPUT_HEADERS.copy()
# plus lookup columns needed explicitly
required_for_lookup = set(required_headers) | {"OLD Item-variant", "Ean no."}
colmap = normalize_columns_case_insensitive(mapping_df, sorted(required_for_lookup))

missing = [name for name in required_for_lookup if colmap.get(name) is None]
if mapping_df.empty:
    st.info("Indlæs en gyldig mapping.xlsx for at fortsætte.")
elif missing:
    st.error(
        "Mapping-filen mangler disse kolonner (matches case-insensitivt): " + ", ".join(missing)
    )
else:
    # Prepare key columns as strings
    old_col = colmap["OLD Item-variant"]
    ean_col = colmap["Ean no."]
    work = mapping_df.copy()
    work[old_col] = work[old_col].astype(str).str.strip()
    work[ean_col] = work[ean_col].astype(str).str.strip()

    # Lookup when user has entered ids
    if not ids:
        st.info("Indsæt værdier for at slå op.")
    else:
        input_df = pd.DataFrame({"input": ids})

        mask = work[old_col].isin(ids) | work[ean_col].isin(ids)
        matches = work.loc[mask].copy()

        # Which inputs did not match
        matched_keys = set(matches[old_col].dropna().astype(str)) | set(matches[ean_col].dropna().astype(str))
        not_found = [x for x in ids if x not in matched_keys]

        # Order output
        ordered = select_and_order(matches, colmap)

        st.subheader("Resultat")
        c1, c2, c3 = st.columns([1, 1, 4])
        with c1:
            st.metric("Antal indlæste", len(ids))
        with c2:
            st.metric("Matches", len(ordered))
        with c3:
            if not_found:
                st.caption("Værdier uden match:")
                st.code("\n".join(not_found), language=None)

        st.dataframe(ordered, use_container_width=True, hide_index=True)

        xlsx = to_excel_bytes(ordered)
        st.download_button(
            label="Download Excel",
            data=xlsx,
            file_name="muuto_mapping_lookup.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# Footnote
st.markdown(
    """
<small>
Bemærk: Formelceller i mapping.xlsx læses som <em>værdier</em> for at sikre at sammenkædede felter i "OLD Item-variant" bliver korrekt brugt til opslag.
</small>
""",
    unsafe_allow_html=True,
)
